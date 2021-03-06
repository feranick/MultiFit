#! /usr/bin/env python

###=============================================================
### Multifit
### Nicola Ferralis <feranick@hotmail.com>
### The entire code is covered by GNU Public License (GPL) v.3
###=============================================================

### Uncomment this if for headless servers.
#import matplotlib
#matplotlib.use('Agg')
### ---------------------------------------
import openpyxl as px
from numpy import *
from lmfit.models import GaussianModel, LorentzianModel, PseudoVoigtModel, VoigtModel
import matplotlib.pyplot as plt
import sys, os.path, getopt, glob
from os.path import exists
from multiprocessing import Pool
import multiprocessing as mp


####################################################################
''' Program definitions and configuration variables '''
####################################################################
class defPar:
    version = '1-20150316a'
    ### Define number of total peaks (do not change: this is read from file)
    NumPeaks = 0
    ### Save results as ASCII?
    ascii = False
    ### Name input paramter file
    inputParFile = 'input_parameters.xlsx'
    if(ascii == True):
        summary = 'summary.txt'        # Save summary fitting results (ASCII)
    else:
        summary = 'summary.xlsx'        # Save summary fitting results (Excel)
    ### Plot initial fitting curve
    initCurve = True
    ### Multiprocessing?
    multiproc = True

####################################################################
''' Main routine to perform and plot the fit '''
####################################################################

def calculate(x, y, x1, y1, ymax, file, type, drawMap, showPlot, lab):
    
    ### Load initialization parameters from xlsx file.
    W = px.load_workbook(defPar.inputParFile, use_iterators = True)
    sheet = W.active
    inval=[]
    defPar.NumPeaks = sheet.get_highest_column() - 1
    
    fpeak = []
    for row in sheet.iter_rows():
        for k in row:
            inval.append(k.internal_value)
    inv = resize(inval, [14, defPar.NumPeaks+1])
    for i in range(1, defPar.NumPeaks+1):
        fpeak.extend([int(inv[1,i])])

	p = Peak(type)
    print (' Fitting with ' + str(defPar.NumPeaks) + ' (' + p.typec + ') peaks')
    ### Initialize parameters for fit.
    pars = p.peak[0].make_params()

    for i in range (0, defPar.NumPeaks):
        if fpeak[i]!=0:
            pars.update(p.peak[i].make_params())
            pars['p{:}_center'.format(str(i))].set(inv[2,i+1], min = inv[3,i+1], max = inv[4,i+1])
            pars['p{:}_sigma'.format(str(i))].set(inv[5,i+1], min = inv[6,i+1], max = inv [7,i+1])
            pars['p{:}_amplitude'.format(str(i))].set(inv[8,i+1]*ymax*80, min=inv[9,i+1], max = inv[10,i+1])
            if (type ==0):
                pars['p{:}_fraction'.format(str(i))].set(inv[11,i+1], min = inv[12,i+1], max = inv[13,i+1])
            if (type ==3):
                pars['p{:}_gamma'.format(str(i))].set(inv[5,i+1])


    ### Add relevant peak to fitting procedure.
    mod = p.peak[0]
    for i in range (1,defPar.NumPeaks):
        if fpeak[i]!=0:
            mod = mod + p.peak[i]

    ### Initialize prefitting curves
    init = mod.eval(pars, x=x)

    ### Perform fitting and display report
    print('\n************************************************************')
    print(' Running fit on file: ' + file + ' (' + str(x1) + ', ' + str(y1) + ')')
    out = mod.fit(y, pars,x=x)
    print(' Done! \n')
    print(' Showing results for: ' + file + ' (' + str(x1) + ', ' + str(y1) + ')')
    print(out.fit_report(min_correl=0.25))

    ### Output file names.
    outfile = 'fit_' + file         # Save individual fitting results
    plotfile = os.path.splitext(file)[0] + '_fit.png'   # Save plot as image

    if os.path.isfile(defPar.summary) == False:
        header = True
    else:
        header = False
        print('\nFit successful: ' + str(out.success))


    if (drawMap == False):
        if (fpeak[1] == 1 & fpeak[2] == 1 & fpeak[5] == 1):
            print('D5/G = {:f}'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude']))
            print('(D4+D5)/G = {:f}'.format((out.best_values['p0_amplitude']+out.best_values['p1_amplitude'])/out.best_values['p5_amplitude']))
            print('D1/G = {:f}'.format(out.best_values['p2_amplitude']/out.best_values['p5_amplitude']))
            if type ==0:
                print('G: {:f}% Gaussian'.format(out.best_values['p5_fraction']*100))
            print('Fit type: {:}'.format(p.typec))
            print('Chi-square: {:}'.format(out.chisqr))
            print('Reduced Chi-square: {:}\n'.format(out.redchi))

            ### Uncomment to enable saving results of each fit in a separate file.
            '''
                with open(outfile, "a") as text_file:
                text_file.write('\nD5/G = {:f}'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude']))
                text_file.write('\n(D4+D5)/G = {:f}'.format((out.best_values['p0_amplitude']+out.best_values['p1_amplitude'])/out.best_values['p5_amplitude']))
                text_file.write('\nD1/G = {:f}'.format(out.best_values['p2_amplitude']/out.best_values['p5_amplitude']))
                if type ==0:
                    text_file.write('\nG %Gaussian: {:f}'.format(out.best_values['p5_fraction']))
                    text_file.write('\nFit type: {:}'.format(p.typec))
                    text_file.write('\nChi-square: {:}'.format(out.chisqr))
                    text_file.write('\nReduced Chi-square: {:}\n'.format(out.redchi))
                '''

    ### Use this for summary in ASCII
    if(defPar.ascii == True):
        with open(defPar.summary, "a") as sum_file:
            if header == True:
                sum_file.write('File\tx1\ty1\tiD1\tiD4\tiD5\tiG\twG\tD5G\t(D4+D5)/G\tD1/G\t%Gaussian\tfit\tChi-square\tred-chi-sq\n')
            sum_file.write('{:}\t'.format(file))
            sum_file.write('{:}\t'.format(x1))
            sum_file.write('{:}\t'.format(y1))
            sum_file.write('{:f}\t'.format(out.best_values['p2_amplitude']))
            sum_file.write('{:f}\t'.format(out.best_values['p0_amplitude']))
            sum_file.write('{:f}\t'.format(out.best_values['p1_amplitude']))
            sum_file.write('{:f}\t'.format(out.best_values['p5_amplitude']))
            sum_file.write('{:f}\t'.format(out.best_values['p5_sigma']*2))
            sum_file.write('{:f}\t'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude']))
            sum_file.write('{:f}\t'.format((out.best_values['p0_amplitude']+out.best_values['p1_amplitude'])/out.best_values['p5_amplitude']))
            sum_file.write('{:f}\t'.format(out.best_values['p2_amplitude']/out.best_values['p5_amplitude']))
            if type ==0:
                sum_file.write('{:f}\t'.format(out.best_values['p1_fraction']))
                sum_file.write('{:f}\t'.format(out.best_values['p2_fraction']))
                sum_file.write('{:f}\t'.format(out.best_values['p5_fraction']))
            else:
                for i in range(0,3):
                    sum_file.write('{:}\t'.format(type-1))
                sum_file.write('{:}\t'.format(p.typec))
            sum_file.write('{:}\t'.format(out.chisqr))
            sum_file.write('{:}\t'.format(out.redchi))
            sum_file.write('{:}\n'.format(lab))
            sum_file.close()

    ### Use this for summary in XLSX
    else:
        if header == True:
            WW=px.Workbook()
            pp=WW.active
            pp.title='Summary'
            summaryHeader = ['File', 'x1', 'y1', 'iD1', 'iD4', 'iD5', 'iG', 'wG', 'D5G', '(D4+D5)/G', \
                 'D1/G', 'D5 %Gaussian','D1 %Gaussian', 'G %Gaussian', 'Fit', \
                 'Chi-square', 'Reduced Chi-square', 'Label']
            pp.append(summaryHeader)
            WW.save(defPar.summary)

        summaryResults = ['{:}'.format(file), float('{:}'.format(x1)), float('{:}'.format(y1)), \
                        float('{:f}'.format(out.best_values['p2_amplitude'])), \
                        float('{:f}'.format(out.best_values['p0_amplitude'])),
                        float('{:f}'.format(out.best_values['p1_amplitude'])), \
                        float('{:f}'.format(out.best_values['p5_amplitude'])), \
                        float('{:f}'.format(out.best_values['p5_sigma']*2)), \
                        float('{:f}'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude'])), \
                        float('{:f}'.format((out.best_values['p0_amplitude']+out.best_values['p1_amplitude'])/out.best_values['p5_amplitude'])), \
                        float('{:f}'.format(out.best_values['p2_amplitude']/out.best_values['p5_amplitude']))]
        if type ==0:
            summaryResults.extend([float('{:f}'.format(out.best_values['p1_fraction'])), \
                                       float('{:f}'.format(out.best_values['p2_fraction'])), \
                                       float('{:f}'.format(out.best_values['p5_fraction']))])
        else:
            summaryResults.extend(['{:}'.format(type-1), '{:}'.format(type-1), '{:}'.format(type-1)])
        summaryResults.extend([p.typec])
        summaryResults.extend([float('{:f}'.format(out.chisqr)), float('{:f}'.format(out.redchi)), lab])


        WW = px.load_workbook(defPar.summary)
        pp = WW.active
        pp.append(summaryResults)
        WW.save(defPar.summary)
        
    if(drawMap == True):
        with open(os.path.splitext(file)[0] + '_map.csv', "a") as map_file:
            map_file.write('{:},'.format(x1))
            map_file.write('{:},'.format(y1))
            map_file.write('{:},'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude']))
            map_file.close()

    else:
        ### Plot optimal fit and individial components
        fig = plt.figure()
        ax = fig.add_subplot(111)
        ax.plot(x, y, label='data')
        if(defPar.initCurve == True):
            ax.plot(x, init, 'k--', label='initial')
        ax.plot(x, out.best_fit, 'r-', label='fit')
        y0 = p.peak[0].eval(x = x, **out.best_values)
        ax.plot(x,y0,'g')
        y = [None]*(defPar.NumPeaks + 1)
        for i in range (1,defPar.NumPeaks):
            if (fpeak[i] ==1):
                y[i] = p.peak[i].eval(x = x, **out.best_values)
                if (i==1 or i==5):
                    ax.plot(x,y[i],'g',linewidth=2.0)
                else:
                    ax.plot(x,y[i],'g')

        ax.text(0.05, 0.875, 'Fit type: {:}\nD5/G = {:f}\nRed. Chi sq: {:}'.format( \
                                p.typec, \
                                out.best_values['p1_amplitude']/out.best_values['p5_amplitude'], \
                                out.redchi), transform=ax.transAxes)

        plt.xlabel('Raman shift [1/cm]')
        plt.ylabel('Intensity [arb. units]')
        plt.title(file)
        plt.legend()
        plt.grid(True)
        plt.savefig(plotfile)  # Save plot
        if(showPlot == True):
            print('*** Close plot to quit ***\n')
            plt.show()
        plt.close()

	del p
	del out


####################################################################
''' Drawing only routine '''
####################################################################

def plotData(x, y, file, showPlot):
    ### Plot initial data
    pngData = os.path.splitext(file)[0] + '.png'   # Save plot as image
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.plot(x, y, label='data')
    plt.xlabel('Raman shift [1/cm]')
    plt.ylabel('Intensity [arb. units]')
    plt.title(file)
    #plt.legend()
    plt.grid(True)
    plt.savefig(pngData)  # Save plot
    if(showPlot == True):
        print('*** Close plot to quit ***\n')
        plt.show()
    plt.close()


####################################################################
''' Definition of class map'''
####################################################################

class Map:
    def __init__(self):
        self.x = []
        self.y = []
        self.z = []
    
    def readCoord(self, file):
        self.num_lines = sum(1 for line in open(file))
        
        data = genfromtxt(file)
        self.x = data[:,0]
        self.y = data[:,1]
        self.z = data[:,2]

    def draw(self, file, showplot):
        self.readCoord(file)
        #fig = plt.figure()
        #ax = fig.add_subplot(111, projection='3d')
        
        #p = ax.pcolor(self.x, self.y, self.z, cmap='Spectral', vmin=min(self.z), vmax=max(self.z))
        #fig.colorbar(p, ax=ax)
        #surf = ax.plot_trisurf(self.x, self.y, self.z, cmap=cm.jet, linewidth=0)
        #fig.colorbar(surf)

        #ax.xaxis.set_major_locator(MaxNLocator(5))
        #ax.yaxis.set_major_locator(MaxNLocator(6))
        #ax.zaxis.set_major_locator(MaxNLocator(5))
        #fig.tight_layout()
        
        #plt.xlabel('[um]')
        #plt.ylabel('[um]')
        #fig.savefig('map.png')  # Save plot
        #if(showplot == True):
        #    print('*** Close plot to quit ***\n')
        #    plt.show()


####################################################################
''' Main program '''
####################################################################

def main():
    print('\n******************************')
    print(' MultiFit v.' + defPar.version)
    print('******************************')

    try:
        opts, args = getopt.getopt(sys.argv[1:], "bftmitph:", ["batch", "file", "type", "map", "input-par", "test", "plot", "help"])
    except getopt.GetoptError:
        # print help information and exit:
        usage()
        sys.exit(2)

    if not exists(defPar.inputParFile):
        print '\n Init parameter not found. Generating a new one...'
        genInitPar()

    if(defPar.multiproc == True):
        print('\n Multiprocessing enabled: ' + str(mp.cpu_count()) + ' CPUs\n')
    else:
        print('\n Multiprocessing disabled\n')
    for o, a in opts:
        if o in ("-b" , "--batch"):
            try:
                type = sys.argv[2]
            except:
                usage()
                sys.exit(2)
            type = int(sys.argv[2])
            i = 0
            if(defPar.multiproc == True):
                p = Pool(mp.cpu_count())
                for f in glob.glob('*.txt'):
                    if (f != 'summary.txt'):
                        rs = readSingleSpectra(f)
                        p.apply_async(calculate, args=(rs.x, rs.y, '0', '0', rs.ymax, f, type, False, False, i))
                        i += 1
                p.close()
                p.join()
            else:
                for f in glob.glob('*.txt'):
                    if (f != 'summary.txt'):
                        rs = readSingleSpectra(f)
                        calculate(rs.x, rs.y, '0', '0', rs.ymax, f, type, False, False, i)
                        i += 1
            addBlankLine(defPar.summary)
        
        elif o in ("-f", "--file"):
            try:
                type = sys.argv[3]
            except:
                usage()
                sys.exit(2)
            file = str(sys.argv[2])
            type = int(sys.argv[3])
            rs = readSingleSpectra(file)
            calculate(rs.x, rs.y, '0', '0', rs.ymax, file, type, False, True, '')


        elif o in ("-p", "--plot"):
            if(len(sys.argv) < 3):
                if(defPar.multiproc == True):
                    p = Pool(mp.cpu_count())
                    for f in glob.glob('*.txt'):
                        if (f != 'summary.txt'):
                            rs = readSingleSpectra(f)
                            print ("Saving plot for: " + f)
                            p.apply_async(plotData, args=(rs.x, rs.y, f, False))
                    p.close()
                    p.join()
                else:
                    for f in glob.glob('*.txt'):
                        if (f != 'summary.txt'):
                            rs = readSingleSpectra(f)
                            print ("Saving plot for: " + f)
                            plotData(rs.x, rs.y, f, False)
            else:
                file = str(sys.argv[2])
                rs = readSingleSpectra(file)
                plotData(rs.x, rs.y, file, True)


        elif o in ("-m", "--map"):
            try:
                type = sys.argv[3]
            except:
                usage()
                sys.exit(2)
            file = str(sys.argv[2])
            type = int(sys.argv[3])
            rm = readMap(file)
            map = Map()
            if(defPar.multiproc == True):
                p = Pool(mp.cpu_count())
                for i in range (1, rm.num_lines):
                    p.apply_async(calculate, args=(rm.x, rm.y[i], rm.x1[i], rm.y1[i], rm.ymax[i], file, type, True, False, ''))
                p.close()
                p.join()
            #map.draw(os.path.splitext(file)[0] + '_map.txt', True)

            else:
                for i in range (1, rm.num_lines):
                    calculate(rm.x, rm.y[i], rm.x1[i], rm.y1[i], rm.ymax[i], file, type, True, False, '')
                #map.draw(os.path.splitext(file)[0] + '_map.txt', True)

        elif o in ("-t", "--test"):
            file = str(sys.argv[2])
            map = Map()
            #map.readCoord(os.path.splitext(file)[0] + '_map.txt')
            map.draw(os.path.splitext(file)[0] + '_map.txt', True)

        elif o in ("-i", "--input-par"):
            genInitPar()

        else:
            usage()
            sys.exit(2)


####################################################################
''' Class to read map files (Horiba LabSpec5) '''
####################################################################

class readMap:

    def __init__(self, file):
        try:
            with open(file) as openfile:
                ### Load data
                self.num_lines = sum(1 for line in openfile)-1
                data = loadtxt(file)
        
                self.x1 = [None]*(self.num_lines)
                self.y1 = [None]*(self.num_lines)
                self.y = [None]*(self.num_lines)
                self.ymax = [None]*(self.num_lines)
    
                self.x = data[0, 2:]
                for i in range(0, self.num_lines):
                    self.x1[i] = data[i+1, 1]
                    self.y1[i] = data[i+1, 0]
                    self.y[i] = data[i+1, 2:]
                    self.ymax[i] = max(self.y[i])
        except:
            print(' File: ' + file + ' not found\n')
            sys.exit(2)


####################################################################
''' Class to read individual spectra '''
####################################################################

class readSingleSpectra:
    def __init__(self, file):
        try:
            with open(file):
                data = loadtxt(file)
                self.x = data[:, 0]
                self.y = data[:, 1]
                self.ymax = max(self.y)
        except:
            print(' File: ' + file + ' not found\n')
            sys.exit(2)


####################################################################
''' Class to define peaks and their properties '''
####################################################################

class Peak:
    ### Define the typology of the peak
    def __init__(self, type):
        
        self.peak= [None]*(defPar.NumPeaks)
        
        if type==0:
            for i in range (0,defPar.NumPeaks):
                self.peak[i] = PseudoVoigtModel(prefix="p"+ str(i) +"_")
            self.typec = "PseudoVoigt"
        elif type == 1:
            for i in range (0,defPar.NumPeaks):
                self.peak[i] = GaussianModel(prefix="p"+ str(i) +"_")
            self.typec = "Gauss"
        elif type == 2:
            for i in range (0,defPar.NumPeaks):
                self.peak[i] = LorentzianModel(prefix="p"+ str(i) +"_")
            self.typec = "Lorentz"
        elif type ==3:
            for i in range (0,defPar.NumPeaks):
                self.peak[i] = VoigtModel(prefix="p"+ str(i) +"_")
            self.typec = "Voigt"
        else:
            print("Warning: type undefined. Using PseudoVoigt")
            for i in range (0,defPar.NumPeaks):
                self.peak[i] = PseudoVoigtModel(prefix="p"+ str(i) +"_")
            self.typec = "PVoigt"

####################################################################
''' Routine to generate initialization parameter file '''
####################################################################

def genInitPar():
    if exists(defPar.inputParFile):
        print(' Input parameter file: ' + defPar.inputParFile + ' already exists\n')
        sys.exit(2)
    else:
        WW=px.Workbook()
        pp=WW.active
        pp.title='InputParameters'
    
        initPar = [['name', 'D4', 'D5', 'D1', 'D3a', 'D3b', 'G', 'D2'], \
               ['activate peak',1,1,1,1,1,1,0], \
               ['center',1160,1260,1330,1400,1500,1590,1680], \
               ['center min','',1240,'','',1500,'',''], \
               ['center max','',1275,'',1440,'','',''], \
               ['sigma',45,45,80,40,40,40,40], \
               ['sigma min',40,40,40,20,20,20,30], \
               ['sigma max',50,50,50,50,50,50,50], \
               ['amplitude',0.1,0.2,1,0.1,0.1,0.8,0.1], \
               ['ampl. min',0,0,0,0,0,0,0], \
               ['ampl. max','','','','','','',''], \
               ['fraction',0.5,0.5,0.5,0.5,0.5,0.5,0.5], \
               ['fraction min',0,0,0,0,0,0,0], \
               ['fraction max',1,1,1,1,1,1,1]]
        for row in range(0, 14):
            pp.append(initPar[row])
    
        WW.save(defPar.inputParFile)
        print(' Input paramters saved in: ' + defPar.inputParFile)


####################################################################
''' Lists the program usage '''
####################################################################

def usage():
    print('Usage: \n\n Single file:')
    print(' python multifit.py -f filename n\n')
    print(' Batch processing:')
    print(' python multifit.py -b n\n')
    print(' Map (acquired with horiba LabSpec5): ')
    print(' python multifit.py -m filename n\n')
    print(' Create and save plot of data only (no fit): ')
    print(' python multifit.py -p filename \n')
    print(' Create and save plot of batch data (no fit): ')
    print(' python multifit.py -p \n')
    print(' Create new input paramter file (xlsx): ')
    print(' python multifit.py -i n\n')
    print(' n = 0: PseudoVoigt 1: Gaussian 2: Lorentzian 3: Voigt\n')


####################################################################
''' Add blank line at the end of the summary spreadsheet '''
####################################################################

def addBlankLine(file):
    if(defPar.ascii == True):
        try:
            with open(file, "a") as sum_file:
                sum_file.write('\n')
        except:
            print "File busy!"
    else:
        WW = px.load_workbook(file)
        pp=WW.active
        pp.append([' '])
        WW.save(file)

####################################################################
''' Main initialization routine '''
####################################################################

if __name__ == "__main__":
    sys.exit(main())
