#!/usr/bin/python

###=============================================================
### Mapfit
### Nicola Ferralis <feranick@hotmail.com>
### The entire code is covered by GNU Public License (GPL) v.3
###=============================================================

### Uncomment this if for headless servers.
#import matplotlib
#matplotlib.use('Agg')
### ---------------------------------------
import openpyxl as px
from numpy import *
from lmfit.models import GaussianModel, LorentzianModel, PseudoVoigtModel
import matplotlib.pyplot as plt
import sys, os.path, getopt, glob
from multiprocessing import Pool
import multiprocessing as mp

class defPar:
    version = '20150218g'
    ### Define number of total peaks
    NumPeaks = 7
    ### Save results as ASCII?
    ascii = False
    ### Multiprocessing?
    multiproc = True


def calculate(x, y, x1, y1, file, type, showplot):
    p = Peak(type)
    fpeak = []
    
    ### Load initialization parameters from xlsx file.
    
    W = px.load_workbook('input_parameters.xlsx', use_iterators = True)
    sheet = W.active
    inval=[]

    for row in sheet.iter_rows():
        for k in row:
            inval.append(k.internal_value)
    inv = resize(inval, [14, defPar.NumPeaks+1])
    for i in range(1, defPar.NumPeaks+1):
        fpeak.extend([int(inv[1,i])])

    ### Initialize parameters for fit.
    pars = p.peak[0].make_params()
    pars['p0_center'].set(inv[2,1], min = inv[3,1], max = inv[4,1] )
    pars['p0_sigma'].set(inv[5,1], min = inv[6,1], max = inv [7,1])
    pars['p0_amplitude'].set(inv[8,1], min=inv[9,1], max = inv[10,1])
    if type ==0:
        pars['p0_fraction'].set(inv[11,1], min = inv[12,1], max = inv[13,1])

    for i in range (1, defPar.NumPeaks):
        if fpeak[i]!=0:
            pars.update(p.peak[i].make_params())
            pars['p{:}_center'.format(str(i))].set(inv[2,i+1], min = inv[3,i+1], max = inv[4,i+1])
            pars['p{:}_sigma'.format(str(i))].set(inv[5,i+1], min = inv[6,i+1], max = inv [7,i+1])
            pars['p{:}_amplitude'.format(str(i))].set(inv[8,i+1], min=inv[9,i+1], max = inv[10,i+1])
            if type ==0:
                pars['p{:}_fraction'.format(str(i))].set(inv[11,i+1], min = inv[12,i+1], max = inv[13,i+1])

    ### Add relevant peak to fitting procedure.
    mod = p.peak[0]
    for i in range (1,defPar.NumPeaks):
        if fpeak[i]!=0:
            mod = mod + p.peak[i]

    ### Initialize and plot initial prefitting curves
    init = mod.eval(pars, x=x)
    fig = plt.figure(1)
    ax = fig.add_subplot(111)
    ax.plot(x, y, label='data')
    ax.plot(x, init, 'k--', label='initial')

    ### Perform fitting and display report
    print('\n************************************************************')
    print(' Running fit on file: ' + file + ' (' + str(x1) + ', ' + str(y1) + ')')
    out = mod.fit(y, pars,x=x)
    print(' Done! \n')
    print(out.fit_report(min_correl=0.25))

    ### Output file names.
    outfile = 'fit_' + file         # Save individual fitting results
    plotfile = os.path.splitext(file)[0] + '_fit.png'   # Save plot as image

    if(defPar.ascii == True):
        summary = 'summary.txt'        # Save summary fitting results (ASCII)
    else:
        summary = 'summary.xlsx'        # Save summary fitting results (Excel)

    if os.path.isfile(summary) == False:
        header = True
    else:
        header = False

        print('\nFit successful: ' + str(out.success))
    if (fpeak[1] == 1 & fpeak[2] == 1 & fpeak[5] == 1):
        print('D5/G = {:f}'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude']))
        print('(D4+D5)/G = {:f}'.format((out.best_values['p0_amplitude']+out.best_values['p1_amplitude'])/out.best_values['p5_amplitude']))
        print('D1/G = {:f}'.format(out.best_values['p2_amplitude']/out.best_values['p5_amplitude']))
        if type ==0:
            print('G: {:f}% Gaussian'.format(out.best_values['p5_fraction']*100))
        print('Fit type: {:}'.format(p.typec))
        print('Chi-square: {:}'.format(out.chisqr))
        print('Reduced Chi-square: {:}\n'.format(out.redchi))

        ### Uncomment to enable saving results of each fit in a separate file.
        '''
        with open(outfile, "a") as text_file:
            text_file.write('\nD5/G = {:f}'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude']))
            text_file.write('\n(D4+D5)/G = {:f}'.format((out.best_values['p0_amplitude']+out.best_values['p1_amplitude'])/out.best_values['p5_amplitude']))
            text_file.write('\nD1/G = {:f}'.format(out.best_values['p2_amplitude']/out.best_values['p5_amplitude']))
            if type ==0:
                text_file.write('\nG %Gaussian: {:f}'.format(out.best_values['p5_fraction']))
            text_file.write('\nFit type: {:}'.format(p.typec))
            text_file.write('\nChi-square: {:}'.format(out.chisqr))
            text_file.write('\nReduced Chi-square: {:}\n'.format(out.redchi))
        '''

        ### Use this for summary in ASCII
        if(defPar.ascii == True):
            with open(summary, "a") as sum_file:
                if header == True:
                    sum_file.write('File\tx1\ty1\tiD1\tiD4\tiD5\tiG\twG\tD5G\t(D4+D5)/G\tD1/G\t%Gaussian\tfit\tChi-square\tred-chi-sq\n')
                sum_file.write('{:}\t'.format(file))
                sum_file.write('{:}\t'.format(x1))
                sum_file.write('{:}\t'.format(y1))
                sum_file.write('{:f}\t'.format(out.best_values['p2_amplitude']))
                sum_file.write('{:f}\t'.format(out.best_values['p0_amplitude']))
                sum_file.write('{:f}\t'.format(out.best_values['p1_amplitude']))
                sum_file.write('{:f}\t'.format(out.best_values['p5_amplitude']))
                sum_file.write('{:f}\t'.format(out.best_values['p5_sigma']*2))
                sum_file.write('{:f}\t'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude']))
                sum_file.write('{:f}\t'.format((out.best_values['p0_amplitude']+out.best_values['p1_amplitude'])/out.best_values['p5_amplitude']))
                sum_file.write('{:f}\t'.format(out.best_values['p2_amplitude']/out.best_values['p5_amplitude']))
                if type ==0:
                    sum_file.write('{:f}\t'.format(out.best_values['p1_fraction']))
                    sum_file.write('{:f}\t'.format(out.best_values['p2_fraction']))
                    sum_file.write('{:f}\t'.format(out.best_values['p5_fraction']))
                else:
                    for i in range(0,3):
                        sum_file.write('{:}\t'.format(type-1))
                sum_file.write('{:}\t'.format(p.typec))
                sum_file.write('{:}\t'.format(out.chisqr))
                sum_file.write('{:}\n'.format(out.redchi))

        ### Use this for summary in XLSX
        else:
            if header == True:
                WW=px.Workbook()
                pp=WW.active
                pp.title='Summary'
                summaryHeader = ['File', 'x1', 'y1', 'iD1', 'iD4', 'iD5', 'iG', 'wG', 'D5G', '(D4+D5)/G', \
                                 'D1/G', 'D5 %Gaussian','D1 %Gaussian', 'G %Gaussian', 'Fit', \
                                 'Chi-square', 'Reduced Chi-square']
                pp.append(summaryHeader)
                WW.save(summary)

            WW = px.load_workbook(summary)
            pp = WW.active

            summaryResults = ['{:}'.format(file), '{:}'.format(x1), '{:}'.format(y1), \
                              '{:f}'.format(out.best_values['p2_amplitude']), \
                              '{:f}'.format(out.best_values['p0_amplitude']),
                              '{:f}'.format(out.best_values['p1_amplitude']), \
                              '{:f}'.format(out.best_values['p5_amplitude']), \
                              '{:f}'.format(out.best_values['p5_sigma']*2), \
                              '{:f}'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude']), \
                              '{:f}'.format((out.best_values['p0_amplitude']+out.best_values['p1_amplitude'])/out.best_values['p5_amplitude']), \
                              '{:f}'.format(out.best_values['p2_amplitude']/out.best_values['p5_amplitude'])]
            if type ==0:
                summaryResults.extend(['{:f}'.format(out.best_values['p1_fraction']), \
                                       '{:f}'.format(out.best_values['p2_fraction']), \
                                       '{:f}'.format(out.best_values['p5_fraction'])])
            else:
                summaryResults.extend(['{:}'.format(type-1), '{:}'.format(type-1), '{:}'.format(type-1)])
            summaryResults.extend([p.typec])
            summaryResults.extend(['{:f}'.format(out.chisqr), '{:f}'.format(out.redchi)])
            pp.append(summaryResults)
            WW.save(summary)

    ### Plot optimal fit and individial components
    ax.plot(x, out.best_fit, 'r-', label='fit')
    y0 = p.peak[0].eval(x = x, **out.best_values)
    ax.plot(x,y0,'g')
    y = [None]*(defPar.NumPeaks + 1)
    for i in range (1,defPar.NumPeaks):
        if (fpeak[i] ==1):
            y[i] = p.peak[i].eval(x = x, **out.best_values)
            if (i==1 or i==5):
                ax.plot(x,y[i],'g',linewidth=2.0)
            else:
                ax.plot(x,y[i],'g')

    ax.text(0.05, 0.9, 'D5/G = {:f}'.format(out.best_values['p1_amplitude']/out.best_values['p5_amplitude']), transform=ax.transAxes)
    ax.text(0.05, 0.9, 'Fit type: {:}\n'.format(p.typec), transform=ax.transAxes)
    
    plt.xlabel('Raman shift [1/cm]')
    plt.ylabel('Intensity [arb. units]')
    plt.legend()
    plt.grid(True)
    plt.savefig(plotfile)
    if(showplot == True):
        print('*** Close plot to quit ***\n')
        plt.show()


###################

def main():
    print('\n******************************')
    print(' MultiFit v.' + defPar.version)
    print('******************************\n')

    try:
        opts, args = getopt.getopt(sys.argv[1:], "bftmh:", ["batch", "file", "type", "map", "help"])
    except getopt.GetoptError:
        # print help information and exit:
        usage()
        sys.exit(2)

    print('Using : ' + str(mp.cpu_count()) + ' CPUs')
    for o, a in opts:
        if o in ("-b" , "--batch"):
            
            type = int(sys.argv[2])
            if(defPar.multiproc == True):
                p = Pool(mp.cpu_count())
                for f in glob.glob('*.txt'):
                    if (f != 'summary.txt'):
                        rs = readSingleSpectra(f)
                        p.apply_async(calculate, args=(rs.x, rs.y, '0', '0', f, type, False))
                p.close()
                p.join()
            else:
                for f in glob.glob('*.txt'):
                    if (f != 'summary.txt'):
                        rs = readSingleSpectra(f)
                        calculate(rs.x, rs.y, '0', '0', f, type, False)
        
        elif o in ("-f", "--file"):
            file = str(sys.argv[2])
            type = int(sys.argv[3])
            rs = readSingleSpectra(file)
            calculate(rs.x, rs.y, '0', '0', file, type, True)

        elif o in ("-m", "--map"):
            file = str(sys.argv[2])
            type = int(sys.argv[3])
            rm = readMap(file)
            if(defPar.multiproc == True):
                p = Pool(mp.cpu_count())
                for i in range (1, rm.num_lines):
                    p.apply_async(calculate, args=(rm.x, rm.y[i], rm.x1[i], rm.y1[i], file, type, False))
                p.close()
                p.join()
            else:
                for i in range (1, rm.num_lines):
                    calculate(rm.x, rm.y[i], rm.x1[i], rm.y1[i], file, type, False)

        else:
            usage()
            sys.exit(2)

class readMap:

    def __init__(self, file):
    ###############################
    
        ### Load data
        self.num_lines = sum(1 for line in open(file))
        data = loadtxt(file)
        
        self.x1 = [None]*self.num_lines
        self.y1 = [None]*self.num_lines
        self.y = [None]*self.num_lines
    
        self.x = data[0, 2:]
        for i in range(1, self.num_lines):
            self.x1[i-1] = data[i, 1]
            self.y1[i-1] = data[i, 2]
            self.y[i-1] = data[i, 2:]

        ###################################

class readSingleSpectra:
    def __init__(self, file):
        data = loadtxt(file)
        self.x = data[:, 0]
        self.y = data[:, 1]

def usage():
    print('Usage: \n\n Single file:')
    print(' python multifit.py -f filename n\n')
    print(' Batch:')
    print(' python multifit.py -b n\n')
    print(' n = 0: PseudoVoigt 1: Gaussian 2: Lorentzian\n')

class Peak:
    ### Define the typology of the peak
    def __init__(self, type):
        
        self.peak= [None]*(defPar.NumPeaks)
        
        if type==0:
            for i in range (0,defPar.NumPeaks):
                self.peak[i] = PseudoVoigtModel(prefix="p"+ str(i) +"_")
            self.typec = "PVoigt"
        else:
            if type == 1:
                for i in range (0,defPar.NumPeaks):
                    self.peak[i] = GaussianModel(prefix="p"+ str(i) +"_")
                self.typec = "Gauss"
            else:
                for i in range (0,defPar.NumPeaks):
                    self.peak[i] = LorentzianModel(prefix="p"+ str(i) +"_")
                self.typec = "Lorentz"

if __name__ == "__main__":
    sys.exit(main())
